from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import StateDB
from .utils import json_loads_or, safe_excel_text, safe_http_url


PRODUCT_HEADERS = [
    "店铺Token", "店铺名称", "店铺链接", "商品名称", "命中关键词",
    "标价", "实际价字段", "库存", "状态", "分类", "商品链接",
    "自动发货", "商品类型", "API域名", "采集时间",
]
SHOP_HEADERS = [
    "店铺Token", "店铺名称", "店铺链接", "命中商品数", "最近读取商品数",
    "状态", "来源评分", "API域名", "发现来源", "发现时间", "最后尝试",
    "最后成功", "连续失败", "下次重试", "错误",
]


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize(ws, fixed: Optional[dict[int, int]] = None) -> None:
    fixed = fixed or {}
    for col in range(1, ws.max_column + 1):
        if col in fixed:
            width = fixed[col]
        else:
            max_len = 0
            for cell in ws[get_column_letter(col)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 80))
            width = min(max(max_len + 2, 10), 52)
        ws.column_dimensions[get_column_letter(col)].width = width


def _safe_cell(value: Any) -> Any:
    if isinstance(value, (int, float)) or value is None:
        return value
    return safe_excel_text(value)


def _add_link(cell, value: Any) -> None:
    url = safe_http_url(value)
    if url:
        cell.hyperlink = url
        cell.style = "Hyperlink"


def _product_row(row: Any) -> list[Any]:
    keywords = "、".join(json_loads_or(row["matched_keywords"], []))
    return [
        row["token"], row["shop_name"], row["shop_url"], row["product_name"], keywords,
        row["listed_price"], row["real_price"], row["stock_count"], row["product_status"],
        row["category_name"], row["product_url"], row["auto_delivery"], row["goods_type"],
        row["api_host"], row["collected_at"],
    ]


def _shop_row(row: Any) -> list[Any]:
    sources = "、".join(json_loads_or(row["sources"], []))
    return [
        row["token"], row["shop_name"], row["shop_url"] or row["url"],
        row["hit_count"], row["scanned_item_count"], row["status"], row["source_score"],
        row["api_host"], sources, row["discovered_at"], row["last_attempt_at"],
        row["last_success_at"], row["consecutive_failures"], row["next_retry_at"], row["last_error"],
    ]


def export_results(db: StateDB, output_dir: Path, prefix: str = "ldxp_gpt_results") -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    candidates, matches, runs = db.rows_for_export()
    status_counts = db.status_counts()
    latest_run = runs[0] if runs else None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"{prefix}_{stamp}.xlsx"
    shops_csv = output_dir / f"{prefix}_shops_{stamp}.csv"
    products_csv = output_dir / f"{prefix}_products_{stamp}.csv"

    wb = Workbook()
    summary = wb.active
    summary.title = "运行摘要"
    summary.append(["指标", "数值"])
    summary_rows = [
        ("导出时间", datetime.now().isoformat(timespec="seconds")),
        ("候选店铺总数", len(candidates)),
        ("当前匹配商品数", len(matches)),
        ("当前命中店铺数", sum(1 for row in candidates if int(row["hit_count"] or 0) > 0)),
        ("成功/有结果", status_counts.get("success", 0)),
        ("部分成功", status_counts.get("partial_success", 0)),
        ("无关键词命中", status_counts.get("no_match", 0)),
        ("空店铺", status_counts.get("empty_shop", 0)),
        ("已关闭", status_counts.get("closed", 0)),
        ("需要验证", status_counts.get("challenge_required", 0)),
        ("被阻断", status_counts.get("blocked", 0)),
        ("限流", status_counts.get("rate_limited", 0)),
        ("网络错误", status_counts.get("network_error", 0)),
        ("接口变化", status_counts.get("api_changed", 0)),
        ("待扫描", status_counts.get("pending", 0)),
    ]
    if latest_run:
        summary_rows.extend(
            [
                ("最近运行开始", latest_run["started_at"]),
                ("最近运行结束", latest_run["finished_at"]),
                ("最近运行尝试数", latest_run["attempted"]),
                ("最近运行成功数", latest_run["successful"]),
                ("最近运行失败数", latest_run["failed"]),
                ("最近运行阻断数", latest_run["blocked"]),
                ("最近运行命中商品数", latest_run["matches"]),
                ("最近运行是否熔断", "是" if latest_run["circuit_broken"] else "否"),
                ("最近运行备注", latest_run["note"]),
            ]
        )
    for key, value in summary_rows:
        summary.append([safe_excel_text(key), _safe_cell(value)])
    _style_header(summary)
    _autosize(summary, {1: 28, 2: 70})

    ws = wb.create_sheet("匹配商品")
    ws.append(PRODUCT_HEADERS)
    for row in matches:
        ws.append([_safe_cell(x) for x in _product_row(row)])
        _add_link(ws.cell(ws.max_row, 3), row["shop_url"])
        _add_link(ws.cell(ws.max_row, 11), row["product_url"])
    _style_header(ws)
    _autosize(ws, {2: 24, 3: 42, 4: 50, 5: 18, 10: 20, 11: 44})

    ws2 = wb.create_sheet("命中店铺")
    ws2.append(SHOP_HEADERS)
    for row in candidates:
        if int(row["hit_count"] or 0) <= 0:
            continue
        ws2.append([_safe_cell(x) for x in _shop_row(row)])
        _add_link(ws2.cell(ws2.max_row, 3), row["shop_url"] or row["url"])
    _style_header(ws2)
    _autosize(ws2, {2: 24, 3: 42, 9: 50, 15: 65})

    ws3 = wb.create_sheet("全部候选")
    ws3.append(SHOP_HEADERS)
    for row in candidates:
        ws3.append([_safe_cell(x) for x in _shop_row(row)])
        _add_link(ws3.cell(ws3.max_row, 3), row["shop_url"] or row["url"])
    _style_header(ws3)
    _autosize(ws3, {2: 24, 3: 42, 9: 50, 15: 65})

    ws4 = wb.create_sheet("失败记录")
    ws4.append(SHOP_HEADERS)
    success_statuses = {"success", "partial_success", "no_match", "empty_shop", "closed", "pending"}
    for row in candidates:
        if row["status"] in success_statuses:
            continue
        ws4.append([_safe_cell(x) for x in _shop_row(row)])
        _add_link(ws4.cell(ws4.max_row, 3), row["shop_url"] or row["url"])
    _style_header(ws4)
    _autosize(ws4, {2: 24, 3: 42, 9: 50, 15: 72})

    ws5 = wb.create_sheet("运行历史")
    run_headers = [
        "运行ID", "开始时间", "结束时间", "命令", "关键词", "引擎", "尝试",
        "成功", "失败", "阻断", "命中商品", "是否熔断", "备注", "配置",
    ]
    ws5.append(run_headers)
    for row in runs:
        ws5.append(
            [
                row["id"], row["started_at"], row["finished_at"], row["command"],
                safe_excel_text("、".join(json_loads_or(row["keywords"], []))), row["engine"],
                row["attempted"], row["successful"], row["failed"], row["blocked"],
                row["matches"], "是" if row["circuit_broken"] else "否",
                safe_excel_text(row["note"]), safe_excel_text(row["config_json"]),
            ]
        )
    _style_header(ws5)
    _autosize(ws5, {5: 28, 13: 55, 14: 65})

    wb.save(xlsx_path)

    with products_csv.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(PRODUCT_HEADERS)
        for row in matches:
            writer.writerow([_safe_cell(x) for x in _product_row(row)])

    with shops_csv.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(SHOP_HEADERS)
        for row in candidates:
            writer.writerow([_safe_cell(x) for x in _shop_row(row)])

    return {"xlsx": xlsx_path, "shops_csv": shops_csv, "products_csv": products_csv}
